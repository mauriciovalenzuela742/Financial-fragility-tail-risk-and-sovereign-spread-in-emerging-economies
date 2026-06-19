"""
extract_mexico.py — Inputs JLoss para México desde el Boletín Estadístico de Banca
Múltiple de la CNBV (Portafolio de Información).

Fuente: portafolioinfo.cnbv.gob.mx -> Boletines -> Banca Múltiple. Archivos por período
  BE_BM_YYYYMM.xls (histórico, BIFF8) / .xlsx (reciente, OOXML). Libro multi-hoja;
  los estados financieros están en la hoja "EF Resumen" (varía: 'EF  Resumen', 'Resumen').

Layout (transpuesto respecto a Perú):
  - Bancos en COLUMNAS (nombres en una fila de encabezado); rubros en la COLUMNA-etiqueta.
  - Una sola columna por banco (consolidado; sin MN/ME/TOTAL).
  - Drift manejado por AUTO-DETECCIÓN: la columna-etiqueta es 0 (boletines viejos) o 1
    (formato CUIF-2022), y aparece una columna agregada "Sistema" que se excluye.
  - La taxonomía detallada del activo cambió con CUIF-2022, pero los rubros AGREGADOS
    (Activo, Pasivo, Capital Contable, Títulos de Crédito Emitidos, Obligaciones
    Subordinadas, Margen financiero, Resultado Neto) conservan el nombre -> ubicación
    por etiqueta es estable.

Partición de deuda (criterio del profesor, bonos vs resto):
  lt_borrow = "Títulos de crédito emitidos" + "Obligaciones subordinadas en circulación"
  st_borrow = Pasivo − lt_borrow

El Estado de Resultados de la CNBV es ACUMULADO YTD -> se des-acumula a flujo mensual.

Outputs (en --out): balance_mexico.csv, coverage_mexico.csv, mktcap_mexico.csv
Uso:
  python download_cnbv.py --start 2010 --end 2026 --out ./cnbv_xls
  python extract_mexico.py --dir ./cnbv_xls --start 2010 --end 2026 --out ./out_mexico
"""
import argparse
import glob
import os
import re
import unicodedata

import numpy as np
import pandas as pd

COUNTRY = "mexico"

# clave panel -> (ticker BMV del grupo | None, pd_source, [alias substring NORMALIZADO]).
# Solo grupos con equity líquido en BMV -> market PD; el resto -> book PD.
# No mapeados NO se descartan: caen a book con slug.
BANKMAP = [
    ("banorte",          "GFNORTEO.MX", "market", ["BANORTE"]),
    ("inbursa",          "GFINBURO.MX", "market", ["INBURSA"]),
    ("banco_bajio",      "BBAJIOO.MX",  "market", ["BAJIO"]),
    ("banregio",         "RA.MX",       "market", ["BANREGIO", "REGIONAL DE MONTERREY"]),
    ("gentera",          "GENTERA.MX",  "market", ["COMPARTAMOS", "GENTERA"]),
    ("bbva_mexico",      None,          "book",   ["BBVA", "BANCOMER"]),
    ("banamex",          None,          "book",   ["BANAMEX", "CITIBANAMEX"]),
    ("santander_mexico", None,          "book",   ["SANTANDER"]),
    ("hsbc_mexico",      None,          "book",   ["HSBC"]),
    ("scotiabank_mexico",None,          "book",   ["SCOTIABANK"]),
    ("mifel",            None,          "book",   ["MIFEL"]),
    ("afirme",           None,          "book",   ["AFIRME"]),
    ("actinver",         None,          "book",   ["ACTINVER"]),
    ("banco_azteca",     None,          "book",   ["AZTECA"]),
    ("bancoppel",        None,          "book",   ["BANCOPPEL"]),
    ("multiva",          None,          "book",   ["MULTIVA"]),
    ("invex",            None,          "book",   ["INVEX"]),
    ("ve_por_mas",       None,          "book",   ["VE POR MAS", "VEPORMAS"]),
    ("monex",            None,          "book",   ["MONEX"]),
    ("intercam",         None,          "book",   ["INTERCAM"]),
    ("cibanco",          None,          "book",   ["CIBANCO"]),
    ("banco_base",       None,          "book",   ["BANCO BASE"]),
    ("autofin",          None,          "book",   ["AUTOFIN"]),
    ("bansi",            None,          "book",   ["BANSI"]),
    ("interacciones",    None,          "book",   ["INTERACCIONES"]),  # histórico (fusionado a Banorte 2018)
    ("ixe",              None,          "book",   ["IXE"]),            # histórico (fusionado a Banorte)
    ("famsa",            None,          "book",   ["FAMSA"]),          # histórico
    ("walmart_banco",    None,          "book",   ["WAL-MART", "WALMART"]),  # histórico
]

# columnas/encabezados que NO son bancos
NON_BANK = {"SISTEMA", "TOTAL", "TOTAL SISTEMA", "BANCA MULTIPLE", "BALANCE GENERAL",
            "ESTADO DE SITUACION FINANCIERA", "ESTADOS FINANCIEROS"}


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().upper()


def _slug(s):
    s = re.sub(r"[^a-z0-9]+", "_", _norm(s).lower()).strip("_")
    return s or "banco_sin_nombre"


def map_bank(name):
    n = _norm(name)
    for key, ticker, src, subs in BANKMAP:
        if any(sub in n for sub in subs):
            return key, ticker, src
    return _slug(name), None, "book"


def period_from_filename(path):
    m = re.search(r"BE_BM_(\d{4})(\d{2})", os.path.basename(path))
    return (int(m.group(1)), int(m.group(2))) if m else None


# ---------------------------------------------------------------------------
# Lector agnóstico al formato (idéntico criterio que Perú)
# ---------------------------------------------------------------------------
def _read_any(path):
    with open(path, "rb") as fh:
        magic = fh.read(8)
    out = {}
    if magic[:4] == b"PK\x03\x04":
        import openpyxl
        with open(path, "rb") as fh:
            wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
            for nm in wb.sheetnames:
                out[nm] = [list(r) for r in wb[nm].iter_rows(values_only=True)]
    elif magic[:4] == b"\xd0\xcf\x11\xe0":
        import xlrd
        wb = xlrd.open_workbook(path)
        for nm in wb.sheet_names():
            sh = wb.sheet_by_name(nm)
            out[nm] = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                       for r in range(sh.nrows)]
    else:
        raise ValueError(f"Formato no reconocido: {path}")
    return out


def _pick_ef_sheet(sheets):
    """Hoja de estados financieros: contiene 'ACTIVO' y 'PASIVO' como rubros y un banco
    conocido en su encabezado. Cubre nombres 'EF Resumen' / 'EF  Resumen' / 'Resumen'."""
    best, best_score = None, -1
    for nm, rows in sheets.items():
        flat = " ".join(_norm(c) for r in rows[:60] for c in r[:3])
        score = ("ACTIVO" in flat) + ("PASIVO" in flat) + ("CAPITAL CONTABLE" in flat) \
                + ("MARGEN FINANCIERO" in flat)
        if "RESUMEN" in nm.upper():
            score += 2
        if score > best_score:
            best, best_score = rows, score
    return best if best_score >= 3 else None


def _detect_grid(rows):
    """Layout transpuesto: (header_row_idx, label_col, [(bankname, col)])."""
    # 1) columna-etiqueta y fila de 'ACTIVO'
    label_col = r_act = None
    for i, r in enumerate(rows[:40]):
        for c in (0, 1, 2):
            if c < len(r) and _norm(r[c]) == "ACTIVO":
                label_col, r_act = c, i
                break
        if label_col is not None:
            break
    if label_col is None:
        return None, None, []
    # 2) fila de encabezado de bancos = la fila (arriba de ACTIVO) con más celdas
    #    de texto a la derecha de label_col
    hdr, best = None, 0
    for i in range(max(0, r_act - 8), r_act):
        cnt = sum(isinstance(rows[i][j], str) and str(rows[i][j]).strip() != ""
                  for j in range(label_col + 1, len(rows[i])))
        if cnt > best:
            best, hdr = cnt, i
    if hdr is None:
        return None, None, []
    # 3) columnas de banco (excluye agregados y celda fecha/unidades)
    banks = []
    for j in range(label_col + 1, len(rows[hdr])):
        name = rows[hdr][j]
        nm = _norm(name)
        if not nm or nm in NON_BANK:
            continue
        if any(t in nm for t in ("MILLONES", "PESOS")) or re.search(r"\b20\d\d\b", nm):
            continue
        banks.append((re.sub(r"\s+", " ", str(name)).strip(), j))
    return hdr, label_col, banks


def _row_index(rows, label_col, *, exact=None, starts=None):
    for i, r in enumerate(rows):
        lab = _norm(r[label_col]) if label_col < len(r) else ""
        if not lab:
            continue
        if exact is not None and lab == exact:
            return i
        if starts is not None and lab.startswith(starts):
            return i
    return None


def _num(v):
    try:
        x = float(v)
        return x if np.isfinite(x) else None
    except (TypeError, ValueError):
        return None


def _parse_per_bank_sheets(sheets, year, month):
    """Formato .xlsm transitorio: una hoja por banco (clave CNBV), con columnas
    'EF del Banco' / subsidiarias / consolidado. Se usa 'EF del Banco' (individual)."""
    recs = []
    for nm, rows in sheets.items():
        r_bg = _row_index(rows, 0, exact="BALANCE GENERAL")
        r_act = _row_index(rows, 0, exact="ACTIVO")
        if r_bg is None or r_act is None:
            continue                                    # no es hoja de banco
        name = next((str(rows[i][0]).strip() for i in range(r_bg - 1, -1, -1)
                     if rows[i] and rows[i][0] and _norm(rows[i][0]) not in
                     ("CONSOLIDACION DE ESTADOS FINANCIEROS", "")), None)
        if not name:
            continue
        # columna de valor = 'EF del Banco' (individual); fallback col 1
        col = 1
        for hr in range(r_act):
            for j, c in enumerate(rows[hr]):
                if _norm(c).startswith("EF DEL BANCO"):
                    col = j
                    break
        key, ticker, src = map_bank(name)
        g = lambda rr: _num(rows[rr][col]) if rr is not None and col < len(rows[rr]) else None
        tot_asset = g(r_act)
        tot_liab  = g(_row_index(rows, 0, exact="PASIVO"))
        equity    = g(_row_index(rows, 0, exact="CAPITAL CONTABLE"))
        tit = g(_row_index(rows, 0, starts="TITULOS DE CREDITO EMITIDOS")) or 0.0
        sub = g(_row_index(rows, 0, starts="OBLIGACIONES SUBORDINADAS EN CIRCULACION")) or 0.0
        lt = tit + sub
        st = (tot_liab - lt) if tot_liab is not None else None
        ni = g(_row_index(rows, 0, exact="RESULTADO NETO"))
        nr = g(_row_index(rows, 0, exact="MARGEN FINANCIERO"))
        if tot_asset is None and tot_liab is None:
            continue
        recs.append({
            "countryname": COUNTRY, "bankname": key, "bankname_cnbv": name,
            "ticker": ticker, "pd_source": src, "year": year, "month": month,
            "date": pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0),
            "tot_asset": tot_asset, "total_liab": tot_liab, "equity": equity,
            "st_borrow": st, "lt_borrow": lt,
            "net_income_ytd": ni, "net_rev_ytd": nr,
        })
    return recs


def parse_file(path):
    per = period_from_filename(path)
    if per is None:
        return []
    year, month = per
    sheets = _read_any(path)
    ef = _pick_ef_sheet(sheets)
    if ef is None:
        return _parse_per_bank_sheets(sheets, year, month)  # formato .xlsm por banco
    hdr, lc, banks = _detect_grid(ef)
    if not banks:
        return []

    r_act = _row_index(ef, lc, exact="ACTIVO")
    r_pas = _row_index(ef, lc, exact="PASIVO")
    r_cap = _row_index(ef, lc, exact="CAPITAL CONTABLE")
    r_tit = _row_index(ef, lc, starts="TITULOS DE CREDITO EMITIDOS")
    r_sub = _row_index(ef, lc, starts="OBLIGACIONES SUBORDINADAS EN CIRCULACION")
    r_mar = _row_index(ef, lc, exact="MARGEN FINANCIERO")
    r_ni  = _row_index(ef, lc, exact="RESULTADO NETO")

    recs = []
    for name, col in banks:
        key, ticker, src = map_bank(name)
        tot_asset = _num(ef[r_act][col]) if r_act is not None else None
        tot_liab  = _num(ef[r_pas][col]) if r_pas is not None else None
        equity    = _num(ef[r_cap][col]) if r_cap is not None else None
        tit = _num(ef[r_tit][col]) if r_tit is not None else 0.0
        sub = _num(ef[r_sub][col]) if r_sub is not None else 0.0
        lt = (tit or 0.0) + (sub or 0.0)
        st = (tot_liab - lt) if tot_liab is not None else None
        ni = _num(ef[r_ni][col]) if r_ni is not None else None
        nr = _num(ef[r_mar][col]) if r_mar is not None else None

        # descarta columnas vacías (banco sin datos ese período)
        if tot_asset is None and tot_liab is None:
            continue
        recs.append({
            "countryname": COUNTRY, "bankname": key, "bankname_cnbv": name,
            "ticker": ticker, "pd_source": src, "year": year, "month": month,
            "date": pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0),
            "tot_asset": tot_asset, "total_liab": tot_liab, "equity": equity,
            "st_borrow": st, "lt_borrow": lt,
            "net_income_ytd": ni, "net_rev_ytd": nr,
        })
    return recs


def decumulate_ytd(df):
    df = df.sort_values(["bankname", "year", "month"]).copy()
    prev_m = df.groupby(["bankname", "year"])["month"].shift(1)
    contiguous = prev_m == (df["month"] - 1)
    first_obs = prev_m.isna()
    for ytd, flow in [("net_income_ytd", "net_income"), ("net_rev_ytd", "net_rev")]:
        prev_v = df.groupby(["bankname", "year"])[ytd].shift(1)
        df[flow] = np.where(contiguous, df[ytd] - prev_v,
                            np.where(first_obs, df[ytd], np.nan))
    df["prof_margin"] = np.where(
        df["net_rev"].fillna(0) != 0, df["net_income"] / df["net_rev"], np.nan)
    return df


def build_balance(file_dir, start, end):
    rows = []
    files = sorted(set(glob.glob(os.path.join(file_dir, "BE_BM_*.xls")) +
                       glob.glob(os.path.join(file_dir, "BE_BM_*.xlsx")) +
                       glob.glob(os.path.join(file_dir, "BE_BM_*.xlsm"))))
    for p in files:
        per = period_from_filename(p)
        if per is None or not (start <= per[0] <= end):
            continue
        try:
            rows += parse_file(p)
        except Exception as e:
            print(f"  [WARN] {os.path.basename(p)}: {e}")
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["_ni"] = df["net_income_ytd"].notna()
    df = (df.sort_values(["bankname", "year", "month", "_ni"])
            .drop_duplicates(["bankname", "year", "month"], keep="last")
            .drop(columns="_ni"))
    df = decumulate_ytd(df)
    cols = ["countryname", "bankname", "bankname_cnbv", "pd_source", "date",
            "year", "month", "tot_asset", "total_liab", "equity",
            "st_borrow", "lt_borrow", "net_income", "net_rev", "prof_margin"]
    return df[cols].sort_values(["bankname", "date"]).reset_index(drop=True)


def coverage_report(df):
    return (df.dropna(subset=["tot_asset"])
              .groupby(["bankname", "pd_source"])
              .agg(n_periodos=("date", "nunique"), desde=("date", "min"),
                   hasta=("date", "max"), nombre_cnbv=("bankname_cnbv", "last"))
              .reset_index().sort_values("desde"))


def fetch_mktcap_yf(df_balance, start, end):
    try:
        import yfinance as yf
    except ImportError:
        print("  [WARN] yfinance no instalado; omito mktcap.")
        return pd.DataFrame(columns=["date", "countryname", "bankname", "price", "mktcap"])
    tmap = {k: t for k, t, s, _ in BANKMAP if s == "market" and t}
    market_banks = df_balance.loc[df_balance.pd_source == "market", "bankname"].unique()
    frames = []
    for bk in market_banks:
        tk = tmap.get(bk)
        if not tk:
            continue
        try:
            t = yf.Ticker(tk)
            hist = t.history(start=f"{start}-01-01", end=f"{end}-12-31", auto_adjust=True)
            if hist.empty:
                continue
            sh = getattr(t, "fast_info", {}).get("shares", None) or t.info.get("sharesOutstanding")
            px = hist["Close"].rename("price").reset_index()
            px["date"] = pd.to_datetime(px["Date"]).dt.tz_localize(None)
            px["countryname"] = COUNTRY
            px["bankname"] = bk
            px["mktcap"] = px["price"] * sh if sh else np.nan
            frames.append(px[["date", "countryname", "bankname", "price", "mktcap"]])
        except Exception as e:
            print(f"  [WARN] mktcap {tk}: {e}")
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
        columns=["date", "countryname", "bankname", "price", "mktcap"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default="./cnbv_xls", help="carpeta con BE_BM_*.xls(x)")
    ap.add_argument("--start", type=int, default=2010)
    ap.add_argument("--end", type=int, default=2026)
    ap.add_argument("--out", default="./out_mexico")
    ap.add_argument("--no-mktcap", action="store_true")
    a = ap.parse_args()

    os.makedirs(a.out, exist_ok=True)
    print("Procesando boletines CNBV (EF Resumen)...")
    bal = build_balance(a.dir, a.start, a.end)
    if bal.empty:
        print("Sin datos. ¿Descargaste los BE_BM_*.xls(x) a --dir?")
        return
    bal.to_csv(os.path.join(a.out, "balance_mexico.csv"), index=False)
    cov = coverage_report(bal)
    cov.to_csv(os.path.join(a.out, "coverage_mexico.csv"), index=False)

    if not a.no_mktcap:
        print("Descargando mktcap (yfinance: grupos BMV)...")
        mkt = fetch_mktcap_yf(bal, a.start, a.end)
        mkt.to_csv(os.path.join(a.out, "mktcap_mexico.csv"), index=False)
        nmk = mkt["bankname"].nunique() if not mkt.empty else 0
    else:
        nmk = "(omitido)"

    print(f"\nbalance_mexico:  {len(bal):>6} filas | {bal['bankname'].nunique()} bancos "
          f"| {bal['date'].min().date()} → {bal['date'].max().date()}")
    print(f"coverage_mexico: {len(cov)} bancos | market={sum(cov['pd_source']=='market')} "
          f"book={sum(cov['pd_source']=='book')}")
    print(f"mktcap_mexico:   {nmk} bancos listados\nOutputs en {a.out}/")


if __name__ == "__main__":
    main()