"""
extract_peru.py — Pipeline integrado de inputs JLoss para Perú (SBS, boletín B-2201).

Cubre los tres formatos históricos de la SBS verificados (2010-2026):
  - BIFF8 (.xls real, OLE2)  -> xlrd          (p.ej. 2010, 2017)
  - OOXML con hojas '1'/'2'  -> openpyxl       (2017+...)
  - OOXML hojas '05-BG (P)'  -> openpyxl       (p.ej. 2015)
Drift manejado por AUTO-DETECCIÓN (no se hardcodea hoja ni columnas):
  - Hoja se elige por título ("Balance General..." / "Estado de Ganancias...").
  - La grilla (fila de subencabezado MN/ME/TOTAL, columnas TOTAL, columna-etiqueta,
    nombres de banco) se detecta por hoja, porque el offset del EGP cambió con los años.

Partición de deuda (criterio del profesor, bonos ≈ 6% del pasivo, reconciliado):
  lt_borrow = OBLIG. EN CIRCULACIÓN NO SUBORDINADAS + SUBORDINADAS   (= bonos emitidos)
  st_borrow = TOTAL PASIVO − lt_borrow                              (residual)

P&L de la SBS es ACUMULADO YTD dentro del año -> se des-acumula a flujo mensual antes
de trimestralizar (de lo contrario σ_ROA del book PD queda sesgada).

Outputs (en --out):
  balance_peru.csv   formato largo v8 por banco-mes
  coverage_peru.csv  cobertura por banco (primer/último período, n)
  mktcap_peru.csv    market cap diario (yfinance) para los bancos con equity listado

Uso:
  python download_sbs.py --code B-2201 --start 2010 --end 2026 --out ./sbs_xlsx
  python extract_peru.py --dir ./sbs_xlsx --start 2010 --end 2026 --out ./out_peru
"""
import argparse
import glob
import os
import re
import unicodedata

import numpy as np
import pandas as pd

COUNTRY = "peru"

# ---------------------------------------------------------------------------
# BANKMAP: clave panel -> {ticker, pd_source, names(substring NORMALIZADO)}.
# 'names' incluye ALIAS HISTÓRICOS (la SBS renombró varias entidades).
# Solo BCP(BAP) e Interbank(IFS) tienen equity listado líquido -> market PD.
# El resto -> book PD contable. Bancos no mapeados NO se descartan: caen a book
# con un slug del nombre SBS (para no perder entidades de la banca múltiple).
# ---------------------------------------------------------------------------
BANKMAP = [
    ("bcp",               "BAP",  "market", ["BANCO DE CREDITO DEL PERU", "BCP"]),
    ("interbank",         "IFS",  "market", ["INTERBANK"]),
    ("bbva_peru",         None,   "book",   ["BBVA", "CONTINENTAL"]),            # ex-Banco Continental
    ("banco_comercio",    None,   "book",   ["BANCOM", "DE COMERCIO"]),
    ("pichincha",         None,   "book",   ["PICHINCHA", "FINANCIERO"]),       # ex-Banco Financiero
    ("banbif",            None,   "book",   ["INTERAMERICANO DE FINANZAS", "BANBIF"]),
    ("scotiabank_peru",   None,   "book",   ["SCOTIABANK", "WIESE", "SUDAMERICANO"]),
    ("citibank_peru",     None,   "book",   ["CITIBANK"]),
    ("mibanco",           None,   "book",   ["MIBANCO"]),
    ("gnb_peru",          None,   "book",   ["GNB"]),
    ("falabella_peru",    None,   "book",   ["FALABELLA"]),
    ("santander_consumer",None,   "book",   ["SANTANDER CONSUMER"]),            # antes que santander_peru
    ("santander_peru",    None,   "book",   ["SANTANDER"]),
    ("ripley_peru",       None,   "book",   ["RIPLEY"]),
    ("alfin_peru",        None,   "book",   ["ALFIN", "AZTECA"]),               # ex-Banco Azteca
    ("icbc_peru",         None,   "book",   ["ICBC"]),
    ("bank_of_china",     None,   "book",   ["BANK OF CHINA", "DE CHINA"]),
    ("bci_peru",          None,   "book",   ["BCI"]),
    ("compartamos",       None,   "book",   ["COMPARTAMOS"]),
    ("hsbc_peru",         None,   "book",   ["HSBC"]),                          # histórico (hasta ~2013)
    ("deutsche_peru",     None,   "book",   ["DEUTSCHE"]),                      # histórico
    ("cencosud_peru",     None,   "book",   ["CENCOSUD"]),                      # histórico (2012-2019)
]

ABBR2MONTH = {"en":1,"fe":2,"ma":3,"ab":4,"my":5,"jn":6,
              "jl":7,"ag":8,"se":9,"oc":10,"no":11,"di":12}


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().upper()


def _slug(s):
    s = _norm(s).lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or "banco_sin_nombre"


def map_bank(sbs_name):
    """(panel_key, ticker, pd_source). No mapeado -> book con slug."""
    n = _norm(sbs_name)
    for key, ticker, src, subs in BANKMAP:
        if any(sub in n for sub in subs):
            return key, ticker, src
    return _slug(sbs_name), None, "book"


# ---------------------------------------------------------------------------
# Lector agnóstico al formato
# ---------------------------------------------------------------------------
def _read_any(path):
    """Devuelve {sheet_name: rows(list[list])}, BIFF8 vía xlrd, OOXML vía openpyxl."""
    with open(path, "rb") as fh:
        magic = fh.read(8)
    out = {}
    if magic[:4] == b"PK\x03\x04":               # OOXML (zip), aunque extensión .XLS
        import openpyxl
        with open(path, "rb") as fh:
            wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
            for nm in wb.sheetnames:
                out[nm] = [list(r) for r in wb[nm].iter_rows(values_only=True)]
    elif magic[:4] == b"\xd0\xcf\x11\xe0":        # OLE2 / BIFF8 (.xls real)
        import xlrd
        wb = xlrd.open_workbook(path)
        for nm in wb.sheet_names():
            sh = wb.sheet_by_name(nm)
            out[nm] = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                       for r in range(sh.nrows)]
    else:
        raise ValueError(f"Formato no reconocido: {path} (magic={magic[:4]!r})")
    return out


def _pick_sheet(sheets, key_substr):
    """Elige la hoja cuyo título (primeras filas) contiene key_substr."""
    for nm, rows in sheets.items():
        for r in rows[:5]:
            if any(key_substr in _norm(c) for c in r):
                return rows
    return None


def _detect_grid(rows):
    """(names_row, label_col, [(bankname, total_col)]) por auto-detección.
    Maneja el offset variable del EGP histórico."""
    h = None
    for i, r in enumerate(rows[:15]):
        if sum(_norm(c) == "TOTAL" for c in r) >= 3:
            h = i
            break
    if h is None:
        return None, None, []
    tcols = [j for j, c in enumerate(rows[h]) if _norm(c) == "TOTAL"]
    label_col = tcols[0] - 3
    names_row = rows[h - 1]
    banks = []
    for tc in tcols:
        name = names_row[tc - 2] if tc - 2 >= 0 else None
        nm = _norm(name)
        if not nm or nm in ("ACTIVO", "PASIVO"):
            continue
        if nm.startswith("TOTAL BANCA") or "SUCURSALES EN EL EXTERIOR" in nm:
            continue                                   # agregados
        banks.append((re.sub(r"\s+", " ", str(name)).strip(), tc))
    return h, label_col, banks


def _row_index(rows, label_col, *, exact=None, starts=None,
               contains_all=None, excludes=None):
    for i, r in enumerate(rows):
        lab = _norm(r[label_col]) if label_col < len(r) else ""
        if not lab:
            continue
        if exact is not None and lab == exact:
            return i
        if starts is not None and lab.startswith(starts):
            return i
        if contains_all is not None and all(t in lab for t in contains_all) \
                and not (excludes and any(t in lab for t in excludes)):
            return i
    return None


def _num(v):
    try:
        x = float(v)
        return x if np.isfinite(x) else None
    except (TypeError, ValueError):
        return None


def period_from_filename(path):
    m = re.search(r"-([a-z]{2})(\d{4})\.", os.path.basename(path).lower())
    if not m:
        return None
    mon = ABBR2MONTH.get(m.group(1))
    return (int(m.group(2)), mon) if mon else None


# ---------------------------------------------------------------------------
# Parser de un archivo (un período)
# ---------------------------------------------------------------------------
def parse_file(path):
    period = period_from_filename(path)
    if period is None:
        return []
    year, month = period
    sheets = _read_any(path)

    bs = _pick_sheet(sheets, "BALANCE GENERAL")
    pl = _pick_sheet(sheets, "GANANCIAS Y P")
    if bs is None:
        return []

    h_b, lc_b, banks_b = _detect_grid(bs)
    r_act = _row_index(bs, lc_b, exact="TOTAL ACTIVO")
    r_pas = _row_index(bs, lc_b, exact="TOTAL PASIVO")
    r_pat = _row_index(bs, lc_b, exact="PATRIMONIO")
    r_no  = _row_index(bs, lc_b, starts="OBLIGACIONES EN CIRCULACION NO SUBORDINADAS")
    r_sub = _row_index(bs, lc_b, starts="OBLIGACIONES EN CIRCULACION SUBORDINADAS")

    # P&L: detectar su propia grilla (offset puede diferir del balance)
    pl_idx = {}
    banks_p_lookup = {}
    if pl is not None:
        h_p, lc_p, banks_p = _detect_grid(pl)
        pl_idx["ni"] = _row_index(pl, lc_p, starts="RESULTADO NETO DEL EJERCICIO")
        pl_idx["nr"] = _row_index(pl, lc_p, starts="MARGEN FINANCIERO BRUTO")
        if pl_idx["ni"] is None:   # plan de cuentas antiguo (p.ej. 2010): "UTILIDAD ( PERDIDA ) NETA"
            pl_idx["ni"] = _row_index(pl, lc_p, contains_all=("UTILIDAD", "NETA"),
                                      excludes=("ANTES",))
        # mapear banco->col en el EGP por nombre (orden/offset pueden diferir)
        for name, tc in banks_p:
            banks_p_lookup[_norm(name)] = tc

    recs = []
    for name, col in banks_b:
        key, ticker, src = map_bank(name)
        tot_asset = _num(bs[r_act][col]) if r_act is not None else None
        tot_liab  = _num(bs[r_pas][col]) if r_pas is not None else None
        equity    = _num(bs[r_pat][col]) if r_pat is not None else None
        bno = _num(bs[r_no][col])  if r_no  is not None else 0.0
        bsu = _num(bs[r_sub][col]) if r_sub is not None else 0.0
        lt = (bno or 0.0) + (bsu or 0.0)
        st = (tot_liab - lt) if tot_liab is not None else None

        ni = nr = None
        if pl is not None:
            pc = banks_p_lookup.get(_norm(name))
            if pc is not None:
                if pl_idx.get("ni") is not None:
                    ni = _num(pl[pl_idx["ni"]][pc])
                if pl_idx.get("nr") is not None:
                    nr = _num(pl[pl_idx["nr"]][pc])

        recs.append({
            "countryname": COUNTRY, "bankname": key, "bankname_sbs": name,
            "ticker": ticker, "pd_source": src,
            "year": year, "month": month,
            "date": pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0),
            "tot_asset": tot_asset, "total_liab": tot_liab, "equity": equity,
            "st_borrow": st, "lt_borrow": lt,
            "net_income_ytd": ni, "net_rev_ytd": nr,
        })
    return recs


# ---------------------------------------------------------------------------
# Des-acumulación YTD del P&L -> flujo mensual
# ---------------------------------------------------------------------------
def decumulate_ytd(df):
    """YTD -> flujo mensual. Solo difiere si el mes previo del mismo año es m-1;
    en enero (o primer mes del año) usa el propio YTD; si hay hueco, deja NaN
    (no inventa flujo). Con la serie mensual completa todos quedan resueltos."""
    df = df.sort_values(["bankname", "year", "month"]).copy()
    prev_m = df.groupby(["bankname", "year"])["month"].shift(1)
    contiguous = prev_m == (df["month"] - 1)
    first_obs = prev_m.isna()
    for ytd, flow in [("net_income_ytd", "net_income"), ("net_rev_ytd", "net_rev")]:
        prev_v = df.groupby(["bankname", "year"])[ytd].shift(1)
        df[flow] = np.where(contiguous, df[ytd] - prev_v,
                            np.where(first_obs, df[ytd], np.nan))
    df["prof_margin"] = np.where(
        df["net_rev"].fillna(0) != 0, df["net_income"] / df["net_rev"], np.nan)
    return df


# ---------------------------------------------------------------------------
# Outputs
# ---------------------------------------------------------------------------
def build_balance(file_dir, start, end):
    rows = []
    files = sorted(glob.glob(os.path.join(file_dir, "B-2201-*.XLS")) +
                   glob.glob(os.path.join(file_dir, "B-2201-*.xls")))
    for p in files:
        per = period_from_filename(p)
        if per is None or not (start <= per[0] <= end):
            continue
        try:
            rows += parse_file(p)
        except Exception as e:
            print(f"  [WARN] {os.path.basename(p)}: {e}")
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # idempotente ante archivos/lecturas duplicadas: conserva la fila con P&L
    df["_ni"] = df["net_income_ytd"].notna()
    df = (df.sort_values(["bankname", "year", "month", "_ni"])
            .drop_duplicates(["bankname", "year", "month"], keep="last")
            .drop(columns="_ni"))
    df = decumulate_ytd(df)
    cols = ["countryname", "bankname", "bankname_sbs", "pd_source", "date",
            "year", "month", "tot_asset", "total_liab", "equity",
            "st_borrow", "lt_borrow", "net_income", "net_rev", "prof_margin"]
    return df[cols].sort_values(["bankname", "date"]).reset_index(drop=True)


def coverage_report(df):
    g = (df.dropna(subset=["tot_asset"])
           .groupby(["bankname", "pd_source"])
           .agg(n_periodos=("date", "nunique"),
                desde=("date", "min"), hasta=("date", "max"),
                nombre_sbs=("bankname_sbs", "last"))
           .reset_index()
           .sort_values("desde"))
    return g


def fetch_mktcap_yf(df_balance, start, end):
    """Market cap diario para bancos con ticker (BAP, IFS). Corre en la máquina del
    usuario (requiere red a Yahoo). Devuelve date, countryname, bankname, price, mktcap."""
    try:
        import yfinance as yf
    except ImportError:
        print("  [WARN] yfinance no instalado; omito mktcap (pip install yfinance).")
        return pd.DataFrame(columns=["date", "countryname", "bankname", "price", "mktcap"])
    tickers = (df_balance[df_balance["pd_source"] == "market"]
               [["bankname"]].drop_duplicates())
    tmap = {"bcp": "BAP", "interbank": "IFS"}
    frames = []
    for _, r in tickers.iterrows():
        bk = r["bankname"]; tk = tmap.get(bk)
        if not tk:
            continue
        try:
            t = yf.Ticker(tk)
            hist = t.history(start=f"{start}-01-01", end=f"{end}-12-31", auto_adjust=True)
            if hist.empty:
                continue
            shares = getattr(t, "fast_info", {}).get("shares", None) or t.info.get("sharesOutstanding")
            px = hist["Close"].rename("price").reset_index()
            px["date"] = pd.to_datetime(px["Date"]).dt.tz_localize(None)
            px["countryname"] = COUNTRY
            px["bankname"] = bk
            px["mktcap"] = px["price"] * shares if shares else np.nan
            frames.append(px[["date", "countryname", "bankname", "price", "mktcap"]])
        except Exception as e:
            print(f"  [WARN] mktcap {tk}: {e}")
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
        columns=["date", "countryname", "bankname", "price", "mktcap"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default="./sbs_xlsx", help="carpeta con B-2201-*.XLS")
    ap.add_argument("--start", type=int, default=2010)
    ap.add_argument("--end", type=int, default=2026)
    ap.add_argument("--out", default="./out_peru")
    ap.add_argument("--no-mktcap", action="store_true")
    a = ap.parse_args()

    os.makedirs(a.out, exist_ok=True)
    print("Procesando balances B-2201...")
    bal = build_balance(a.dir, a.start, a.end)
    if bal.empty:
        print("Sin datos. ¿Descargaste con download_sbs.py a --dir?")
        return
    bal.to_csv(os.path.join(a.out, "balance_peru.csv"), index=False)

    cov = coverage_report(bal)
    cov.to_csv(os.path.join(a.out, "coverage_peru.csv"), index=False)

    if not a.no_mktcap:
        print("Descargando mktcap (yfinance: BAP, IFS)...")
        mkt = fetch_mktcap_yf(bal, a.start, a.end)
        mkt.to_csv(os.path.join(a.out, "mktcap_peru.csv"), index=False)
        nmk = mkt["bankname"].nunique() if not mkt.empty else 0
    else:
        nmk = "(omitido)"

    print(f"\nbalance_peru:  {len(bal):>6} filas | {bal['bankname'].nunique()} bancos "
          f"| {bal['date'].min().date()} → {bal['date'].max().date()}")
    print(f"coverage_peru: {len(cov)} bancos | market={sum(cov['pd_source']=='market')} "
          f"book={sum(cov['pd_source']=='book')}")
    print(f"mktcap_peru:   {nmk} bancos listados")
    print(f"\nOutputs en {a.out}/")


if __name__ == "__main__":
    main()